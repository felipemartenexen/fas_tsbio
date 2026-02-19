#!/usr/bin/env python3
"""
Gerador de PDF: Documentação dos Indicadores TSBio
Territórios Sustentáveis da Bioeconomia

Uso:
    pip install openpyxl reportlab
    python gerar_documentacao_tsbio.py

Requisitos:
    - Python 3.8+
    - openpyxl >= 3.1
    - reportlab >= 4.0
    - Arquivo '_documentacao.xlsx' na mesma pasta do script
"""

import os
import sys
from collections import OrderedDict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Erro: instale openpyxl → pip install openpyxl")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, PageBreak,
        Table, TableStyle, KeepTogether, Flowable, Frame,
        PageTemplate, BaseDocTemplate, NextPageTemplate
    )
except ImportError:
    sys.exit("Erro: instale reportlab → pip install reportlab")


# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES
# ═══════════════════════════════════════════════════════════════════════════

# Caminho do arquivo Excel de entrada (ajuste se necessário)
EXCEL_INPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), r"C:\Users\luiz.felipe\Desktop\FLP\MapiaEng\GitHub\fas_tsbio\data\Indicadores_processado_por_tema\outputs\_documentacao.xlsx")

# Caminho do PDF de saída
PDF_OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "documentacao_indicadores_tsbio.pdf")

# Máximo de variáveis exibidas por indicador antes de truncar
MAX_VARS_DISPLAY = 20

# Data exibida no cabeçalho
_MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}
_now = datetime.now()
MES_ANO = f"{_MESES_PT[_now.month]} {_now.year}"


# ═══════════════════════════════════════════════════════════════════════════
# DESCRIÇÕES DAS CATEGORIAS
# ═══════════════════════════════════════════════════════════════════════════

CATEGORY_DESCRIPTIONS = {
    "Agropecuária": (
        "Reúne indicadores sobre a estrutura produtiva agropecuária dos municípios, "
        "incluindo perfil dos estabelecimentos, tipo de atividade (lavoura, pecuária, "
        "aquicultura, pesca, produção florestal), mecanização, uso de insumos, assistência "
        "técnica, agricultura familiar e produção animal e vegetal. As principais fontes "
        "são o Censo Agropecuário 2017, a Pesquisa Agrícola Municipal (PAM), a Pesquisa "
        "Pecuária Municipal (PPM) e a MUNIC — todas do IBGE."
    ),
    "Ambiental": (
        "Indicadores de monitoramento ambiental dos territórios, abrangendo uso e "
        "cobertura do solo, desmatamento, degradação florestal e áreas queimadas. "
        "Os dados são provenientes de sistemas de referência como MapBiomas, PRODES "
        "e DETER, que utilizam sensoriamento remoto para acompanhar as transformações "
        "na paisagem ao longo do tempo."
    ),
    "Cooperativa": (
        "Dados sobre o cooperativismo nos municípios, com foco em cooperativas de "
        "crédito e número de cooperados. As fontes incluem o Banco Central do Brasil "
        "e o Desafio Conexus, permitindo avaliar o grau de organização financeira "
        "cooperativa nas comunidades locais."
    ),
    "Domicílios": (
        "Indicadores sobre as condições dos domicílios particulares permanentes, "
        "incluindo acesso a serviços básicos como abastecimento de água, esgotamento "
        "sanitário, coleta de lixo, energia elétrica e internet. Também contempla "
        "características construtivas e posse de bens. Os dados são oriundos do "
        "Censo Demográfico 2022 do IBGE."
    ),
    "Economia": (
        "Indicador do Produto Interno Bruto (PIB) municipal, que representa o valor "
        "total dos bens e serviços produzidos no município. Permite dimensionar a "
        "atividade econômica local e comparar o desempenho entre os territórios. "
        "Fonte: IBGE — Contas Nacionais."
    ),
    "Educação": (
        "Indicadores educacionais que incluem taxas de alfabetização, nível de "
        "instrução, número médio de anos de estudo e frequência escolar, "
        "desagregados por sexo, cor ou raça e faixa etária. Os dados permitem "
        "analisar desigualdades educacionais nos territórios. Fonte: Censo "
        "Demográfico 2022."
    ),
    "Entorno Domicílios": (
        "Caracterização da infraestrutura do entorno dos domicílios urbanos, "
        "incluindo presença de arborização, calçada, iluminação pública, "
        "bueiros, meio-fio, rampa para cadeirantes, identificação do logradouro "
        "e pavimentação. Esses dados refletem a qualidade do espaço urbano. "
        "Fonte: Censo Demográfico 2022."
    ),
    "Extrativismo": (
        "Dados sobre a produção extrativista vegetal e silvicultura nos municípios, "
        "incluindo volumes e valores da produção de produtos florestais madeireiros "
        "e não madeireiros. Fonte: Pesquisa da Extração Vegetal e da Silvicultura "
        "(PEVS) do IBGE."
    ),
    "Favelas e Comunidades Urbanas": (
        "Indicadores específicos sobre as condições de vida em favelas e comunidades "
        "urbanas, abrangendo características dos domicílios, acesso a serviços de "
        "saneamento, condições do entorno, perfil demográfico (cor/raça, sexo, idade) "
        "e escolaridade da população residente. Fonte: Censo Demográfico 2022."
    ),
    "Fundiário e Áreas Protegidas": (
        "Informações sobre a estrutura fundiária e as áreas legalmente protegidas "
        "nos territórios, incluindo malha fundiária, terras indígenas, unidades de "
        "conservação, assentamentos federais, territórios quilombolas e Cadastro "
        "Ambiental Rural (CAR). As fontes incluem ICMBio, FUNAI, INCRA, SICAR "
        "e Cartas da Terra."
    ),
    "Indígenas": (
        "Conjunto abrangente de indicadores sobre a população indígena, incluindo "
        "demografia, alfabetização, condições de moradia, acesso a serviços básicos, "
        "distribuição por etnia e língua falada, situação urbana ou rural e registro "
        "de nascimento. Os dados permitem caracterizar as condições de vida dos povos "
        "indígenas nos territórios. Fonte: Censo Demográfico 2022."
    ),
    "Infraestrutura": (
        "Indicadores sobre a infraestrutura municipal de saneamento básico, "
        "informática e comunicação, e transporte e mobilidade urbana. As fontes "
        "incluem a MUNIC/IBGE e o Instituto Água e Saneamento (IAS), abrangendo "
        "a capacidade instalada dos municípios em serviços essenciais."
    ),
    "Políticas Públicas": (
        "Indicadores sobre a presença e alcance de políticas públicas nos territórios, "
        "incluindo Bolsa Família, PNAE (alimentação escolar), PAA (aquisição de alimentos), "
        "PGPM-Bio (sociobiodiversidade), crédito rural (PRONAF/SICOR), produção orgânica, "
        "economia solidária, Cadastro Único e governança municipal. As fontes são "
        "diversas: SAGI, CONAB, MAPA, FNDE, IBGE MUNIC, entre outras."
    ),
    "População": (
        "Indicadores demográficos dos municípios, incluindo população total, "
        "densidade demográfica, distribuição por sexo, cor ou raça, faixa etária, "
        "razão de dependência, taxa de fecundidade, população em situação urbana e "
        "rural, e registro de nascimento. As fontes são o Censo Demográfico 2022 "
        "e estimativas do IBGE."
    ),
    "Quilombola": (
        "Indicadores detalhados sobre a população quilombola, abrangendo "
        "demografia, alfabetização, condições de moradia, acesso a saneamento e "
        "energia, características do entorno, posse de bens, nível de instrução, "
        "registro de nascimento e distribuição territorial. Os dados são essenciais "
        "para políticas voltadas a comunidades remanescentes de quilombos. "
        "Fonte: Censo Demográfico 2022."
    ),
    "Religião": (
        "Indicadores sobre a diversidade religiosa nos territórios, incluindo "
        "distribuição da população por grandes grupos de religião, cruzamentos "
        "com cor ou raça, sexo e taxa de alfabetização. Fonte: Censo Demográfico 2022."
    ),
    "Trabalho e Renda": (
        "Indicadores sobre o mercado de trabalho municipal, incluindo formalização "
        "(carteira assinada, CNPJ), distribuição entre setor público e privado, "
        "número de trabalhos exercidos, rendimento por cor ou raça e posição na "
        "ocupação. Fonte: Censo Demográfico 2022."
    ),
    "Vulnerabiliade Saúde": (
        "Indicadores de vulnerabilidade climática aplicados à saúde pública, oriundos "
        "da plataforma Adapta Brasil (MCTI). Abrangem riscos relacionados a arboviroses "
        "(dengue, zika, chikungunya), doenças respiratórias e outros agravos sensíveis "
        "ao clima. Incluem ameaça climática, exposição, sensibilidade, capacidade "
        "adaptativa e vulnerabilidade, com cenários para 2030 e 2050 (otimista e pessimista)."
    ),
    "Vulnerabilidade Biodiversidade": (
        "Indicadores de vulnerabilidade da biodiversidade frente às mudanças climáticas, "
        "da plataforma Adapta Brasil (MCTI). Avaliam a integridade dos biomas considerando "
        "ameaça climática, uso de agrotóxicos, áreas protegidas, desmatamento, "
        "fragmentação e restauração de ecossistemas, com projeções para diferentes "
        "cenários de aquecimento global (SWL 1.0 e 2.0)."
    ),
    "Vulnerabilidade Desastres Geo-hidrológicos": (
        "Indicadores de vulnerabilidade a desastres geo-hidrológicos, incluindo "
        "deslizamentos de terra e inundações, da plataforma Adapta Brasil (MCTI). "
        "Contemplam ameaça climática, exposição, sensibilidade, capacidade adaptativa, "
        "ações de redução de risco e cenários futuros (2030 e 2050, otimista e pessimista). "
        "Essenciais para planejamento de defesa civil e resiliência territorial."
    ),
    "Vulnerabilidade Recursos Hídricos": (
        "Indicadores de vulnerabilidade dos recursos hídricos, da plataforma "
        "Adapta Brasil (MCTI). Avaliam o risco de estresse hídrico nos municípios, "
        "considerando ameaça de escassez, demanda de água, capacidade dos reservatórios, "
        "ações de prevenção e alternativas de abastecimento, com projeções para "
        "cenários futuros (2030 e 2050)."
    ),
    "Vulnerabilidade Segurança Alimentar": (
        "Indicadores de vulnerabilidade da segurança alimentar frente às mudanças "
        "climáticas, da plataforma Adapta Brasil (MCTI). Abrangem acesso e consumo "
        "de alimentos, disponibilidade e produção agropecuária, estabilidade do "
        "abastecimento e utilização biológica, com cenários futuros (2030 e 2050). "
        "Permitem identificar municípios em situação de insegurança alimentar."
    ),
    "Vulnerabilidade Segurança Energética": (
        "Indicadores de vulnerabilidade da segurança energética, da plataforma "
        "Adapta Brasil (MCTI). Avaliam o acesso à energia, disponibilidade de "
        "fontes renováveis (solar, eólica, hidrelétrica), demanda de resfriamento, "
        "pobreza energética e capacidade adaptativa dos municípios, com projeções "
        "para cenários de aquecimento global (SWL 2.0)."
    ),
    "Índices": (
        "Índices sintéticos de desenvolvimento e desigualdade, incluindo o Índice "
        "de Gini (desigualdade de renda, de 0 a 1) e o Índice de Desenvolvimento "
        "Humano Municipal (IDHM), composto pelas dimensões longevidade, educação e "
        "renda. Fontes: Atlas do Desenvolvimento Humano no Brasil."
    ),
}


# ═══════════════════════════════════════════════════════════════════════════
# PALETA DE CORES
# ═══════════════════════════════════════════════════════════════════════════

PRIMARY         = HexColor("#1B5E20")   # Verde escuro principal
PRIMARY_LIGHT   = HexColor("#E8F5E9")   # Verde claro fundo
ACCENT          = HexColor("#2E7D32")   # Verde médio
ACCENT2         = HexColor("#43A047")   # Verde claro
GRAY_DARK       = HexColor("#333333")
GRAY_MED        = HexColor("#666666")
GRAY_LIGHT      = HexColor("#999999")
GRAY_BG         = HexColor("#F5F5F5")
BORDER_COLOR    = HexColor("#C8E6C9")
WHITE           = white

PAGE_W, PAGE_H  = A4
MARGIN_LEFT     = 20 * mm
MARGIN_RIGHT    = 20 * mm
MARGIN_TOP      = 20 * mm
MARGIN_BOTTOM   = 25 * mm
CONTENT_WIDTH   = PAGE_W - MARGIN_LEFT - MARGIN_RIGHT


# ═══════════════════════════════════════════════════════════════════════════
# FLOWABLES CUSTOMIZADOS
# ═══════════════════════════════════════════════════════════════════════════

class ColorBar(Flowable):
    """Barra colorida horizontal."""
    def __init__(self, width, height, color):
        Flowable.__init__(self)
        self.width = width
        self.height = height
        self.color = color

    def draw(self):
        self.canv.setFillColor(self.color)
        self.canv.roundRect(0, 0, self.width, self.height, 2, fill=1, stroke=0)


class HLine(Flowable):
    """Linha horizontal separadora."""
    def __init__(self, width, color=BORDER_COLOR, thickness=0.5):
        Flowable.__init__(self)
        self.width = width
        self.color = color
        self.thickness = thickness

    def draw(self):
        self.canv.setStrokeColor(self.color)
        self.canv.setLineWidth(self.thickness)
        self.canv.line(0, 0, self.width, 0)


class RoundedBox(Flowable):
    """Caixa com fundo colorido e cantos arredondados contendo um Paragraph."""
    def __init__(self, text_paragraph, width, bg_color, padding=3*mm, radius=3):
        Flowable.__init__(self)
        self.text_paragraph = text_paragraph
        self.box_width = width
        self.bg_color = bg_color
        self.padding = padding
        self.radius = radius
        # Calcula altura do texto
        tw, th = text_paragraph.wrap(width - 2 * padding, 10000)
        self.width = width
        self.height = th + 2 * padding

    def draw(self):
        self.canv.setFillColor(self.bg_color)
        self.canv.roundRect(
            0, 0, self.box_width, self.height,
            self.radius, fill=1, stroke=0
        )
        self.text_paragraph.drawOn(
            self.canv, self.padding, self.padding
        )


# ═══════════════════════════════════════════════════════════════════════════
# ESTILOS DE TEXTO
# ═══════════════════════════════════════════════════════════════════════════

styles = getSampleStyleSheet()

style_cover_title = ParagraphStyle(
    'CoverTitle', parent=styles['Title'],
    fontSize=32, leading=38, textColor=PRIMARY,
    fontName='Helvetica-Bold', alignment=TA_LEFT,
    spaceAfter=6*mm,
)
style_cover_tsb = ParagraphStyle(
    'CoverTSB', parent=styles['Title'],
    fontSize=42, leading=50, textColor=ACCENT,
    fontName='Helvetica-Bold', alignment=TA_LEFT,
    spaceAfter=4*mm,
)
style_cover_sub = ParagraphStyle(
    'CoverSub', parent=styles['Normal'],
    fontSize=16, leading=22, textColor=ACCENT,
    fontName='Helvetica', alignment=TA_LEFT,
    spaceAfter=4*mm,
)
style_cover_stats = ParagraphStyle(
    'CoverStats', parent=styles['Normal'],
    fontSize=13, leading=18, textColor=GRAY_MED,
    fontName='Helvetica', alignment=TA_LEFT,
)
style_toc_title = ParagraphStyle(
    'TocTitle', parent=styles['Title'],
    fontSize=22, leading=28, textColor=PRIMARY,
    fontName='Helvetica-Bold', alignment=TA_LEFT,
    spaceAfter=8*mm,
)
style_toc_item = ParagraphStyle(
    'TocItem', parent=styles['Normal'],
    fontSize=10.5, leading=19, textColor=GRAY_DARK,
    fontName='Helvetica', alignment=TA_LEFT,
    leftIndent=6*mm,
)
style_cat_desc = ParagraphStyle(
    'CatDesc', parent=styles['Normal'],
    fontSize=9, leading=13.5, textColor=GRAY_MED,
    fontName='Helvetica', alignment=TA_JUSTIFY,
    spaceAfter=4*mm, spaceBefore=2*mm,
    leftIndent=2*mm, rightIndent=2*mm,
)
style_ind_title = ParagraphStyle(
    'IndTitle', parent=styles['Heading2'],
    fontSize=11, leading=14.5, textColor=PRIMARY,
    fontName='Helvetica-Bold', alignment=TA_LEFT,
    spaceAfter=1.5*mm, spaceBefore=0,
)
style_ind_desc = ParagraphStyle(
    'IndDesc', parent=styles['Normal'],
    fontSize=8.8, leading=12.5, textColor=GRAY_DARK,
    fontName='Helvetica-Oblique', alignment=TA_LEFT,
    spaceAfter=2*mm, leftIndent=0,
)
style_ind_meta = ParagraphStyle(
    'IndMeta', parent=styles['Normal'],
    fontSize=8.5, leading=12, textColor=GRAY_DARK,
    fontName='Helvetica', alignment=TA_LEFT,
    spaceAfter=1.5*mm,
)
style_var_label = ParagraphStyle(
    'VarLabel', parent=styles['Normal'],
    fontSize=8.5, leading=12, textColor=GRAY_DARK,
    fontName='Helvetica-Bold', alignment=TA_LEFT,
    spaceAfter=1*mm,
)
style_var_text = ParagraphStyle(
    'VarText', parent=styles['Normal'],
    fontSize=7.8, leading=11, textColor=GRAY_MED,
    fontName='Courier', alignment=TA_LEFT,
    spaceAfter=0,
)


# ═══════════════════════════════════════════════════════════════════════════
# TEMPLATE DO DOCUMENTO
# ═══════════════════════════════════════════════════════════════════════════

class TSBioDocTemplate(BaseDocTemplate):
    """Template customizado com capa e páginas de conteúdo."""

    def __init__(self, filename, **kwargs):
        BaseDocTemplate.__init__(self, filename, **kwargs)
        frame = Frame(
            MARGIN_LEFT, MARGIN_BOTTOM,
            CONTENT_WIDTH, PAGE_H - MARGIN_TOP - MARGIN_BOTTOM,
            id='normal'
        )
        self.addPageTemplates([
            PageTemplate(id='cover', frames=[frame], onPage=self._cover_page),
            PageTemplate(id='content', frames=[frame], onPage=self._content_page),
        ])

    def _cover_page(self, canvas, doc):
        canvas.saveState()
        # Barra verde no topo
        canvas.setFillColor(PRIMARY)
        canvas.rect(0, PAGE_H - 8*mm, PAGE_W, 8*mm, fill=1, stroke=0)
        # Barra verde na base
        canvas.setFillColor(ACCENT)
        canvas.rect(0, 0, PAGE_W, 4*mm, fill=1, stroke=0)
        canvas.restoreState()

    def _content_page(self, canvas, doc):
        canvas.saveState()
        # Linha do cabeçalho
        canvas.setStrokeColor(BORDER_COLOR)
        canvas.setLineWidth(0.5)
        canvas.line(MARGIN_LEFT, PAGE_H - 12*mm, PAGE_W - MARGIN_RIGHT, PAGE_H - 12*mm)
        # Texto do cabeçalho
        canvas.setFont('Helvetica', 7.5)
        canvas.setFillColor(GRAY_LIGHT)
        canvas.drawString(MARGIN_LEFT, PAGE_H - 10*mm, "Documentação dos Indicadores TSBio")
        canvas.drawRightString(PAGE_W - MARGIN_RIGHT, PAGE_H - 10*mm, MES_ANO)
        # Linha do rodapé
        canvas.setStrokeColor(BORDER_COLOR)
        canvas.line(MARGIN_LEFT, 18*mm, PAGE_W - MARGIN_RIGHT, 18*mm)
        # Número da página
        canvas.setFont('Helvetica', 7.5)
        canvas.setFillColor(GRAY_LIGHT)
        canvas.drawCentredString(PAGE_W / 2, 13*mm, f"Página {doc.page}")
        # Barra verde fina na base
        canvas.setFillColor(PRIMARY)
        canvas.rect(0, 0, PAGE_W, 2*mm, fill=1, stroke=0)
        canvas.restoreState()


# ═══════════════════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ═══════════════════════════════════════════════════════════════════════════

def load_data(excel_path):
    """Carrega os indicadores do arquivo Excel."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['indicadores']

    indicators = []
    for r in range(2, ws.max_row + 1):
        cat    = ws.cell(r, 1).value
        fonte  = ws.cell(r, 2).value
        tema   = ws.cell(r, 3).value
        desc   = ws.cell(r, 4).value
        linhas = ws.cell(r, 5).value
        n_cols = ws.cell(r, 6).value
        cols   = ws.cell(r, 7).value

        if not cat or not tema:
            continue

        variables = []
        if cols:
            variables = [v.strip() for v in str(cols).split(';') if v.strip()]

        indicators.append({
            'categoria':  cat,
            'fonte':      fonte or '',
            'tema':       tema,
            'descricao':  desc or '',
            'registros':  linhas or 0,
            'n_colunas':  n_cols or 0,
            'variaveis':  variables,
        })

    # Agrupar por categoria preservando ordem
    categories = OrderedDict()
    for ind in indicators:
        cat = ind['categoria']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(ind)

    return categories, len(indicators)


def escape_xml(text):
    """Escapa caracteres especiais para uso em Paragraph do ReportLab."""
    if not text:
        return ""
    return (text
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;'))


def build_category_banner(cat_name, count):
    """Cria o banner verde do cabeçalho de categoria."""
    title_p = Paragraph(
        f'<font size="18"><b>{escape_xml(cat_name)}</b></font><br/>'
        f'<font size="9" color="#A5D6A7">{count} indicador{"es" if count != 1 else ""} nesta categoria</font>',
        ParagraphStyle(
            'BannerText', parent=styles['Normal'],
            textColor=WHITE, fontName='Helvetica-Bold',
            fontSize=18, leading=24, alignment=TA_LEFT,
        )
    )
    banner = Table([[title_p]], colWidths=[CONTENT_WIDTH - 10*mm])
    banner.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), PRIMARY),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5*mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5*mm),
        ('TOPPADDING', (0, 0), (-1, -1), 4*mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4*mm),
        ('ROUNDEDCORNERS', [4, 4, 4, 4]),
    ]))
    return banner


def build_category_description(cat_name):
    """Retorna o Paragraph com a descrição da categoria, se disponível."""
    desc_text = CATEGORY_DESCRIPTIONS.get(cat_name)
    if not desc_text:
        return None

    desc_p = Paragraph(escape_xml(desc_text), style_cat_desc)
    # Caixa com fundo cinza claro
    box = RoundedBox(desc_p, CONTENT_WIDTH, GRAY_BG, padding=3.5*mm, radius=3)
    return box


def build_indicator_block(ind):
    """Constrói o bloco de um indicador individual."""
    elements = []
    n_vars = len(ind['variaveis'])

    # Título do indicador
    elements.append(Paragraph(escape_xml(ind['tema']), style_ind_title))

    # Descrição
    if ind['descricao']:
        elements.append(Paragraph(escape_xml(ind['descricao']), style_ind_desc))

    # Fonte e Registros (sem status)
    fonte_hex = ACCENT.hexval()
    meta_text = (
        f'<font color="{fonte_hex}"><b>Fonte:</b></font> {escape_xml(ind["fonte"])} '
        f'<font color="#999999">|</font> '
        f'<font color="{fonte_hex}"><b>Registros:</b></font> {ind["registros"]}'
    )
    elements.append(Paragraph(meta_text, style_ind_meta))

    # Variáveis
    if n_vars > 0:
        elements.append(Paragraph(f'<b>Variáveis ({n_vars}):</b>', style_var_label))
        if n_vars <= MAX_VARS_DISPLAY:
            vars_text = ', '.join(ind['variaveis'])
        else:
            vars_text = ', '.join(ind['variaveis'][:MAX_VARS_DISPLAY])
            vars_text += f' ... (+{n_vars - MAX_VARS_DISPLAY} variáveis)'
        elements.append(Paragraph(escape_xml(vars_text), style_var_text))

    # Separador
    elements.append(Spacer(1, 2*mm))
    elements.append(HLine(CONTENT_WIDTH))
    elements.append(Spacer(1, 3*mm))

    return KeepTogether(elements)


# ═══════════════════════════════════════════════════════════════════════════
# CONSTRUÇÃO DO PDF
# ═══════════════════════════════════════════════════════════════════════════

def generate_pdf(excel_path, output_path):
    """Função principal que gera o PDF completo."""
    print(f"Carregando dados de: {excel_path}")
    categories, total_indicators = load_data(excel_path)
    total_categories = len(categories)
    print(f"  → {total_indicators} indicadores em {total_categories} categorias")

    doc = TSBioDocTemplate(
        output_path,
        pagesize=A4,
        topMargin=MARGIN_TOP,
        bottomMargin=MARGIN_BOTTOM,
        leftMargin=MARGIN_LEFT,
        rightMargin=MARGIN_RIGHT,
    )

    story = []

    # ── CAPA ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 50*mm))
    story.append(Paragraph("Documentação dos Indicadores", style_cover_title))
    story.append(Paragraph("TSBio", style_cover_tsb))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("Territórios Sustentáveis da Bioeconomia", style_cover_sub))
    story.append(Spacer(1, 12*mm))

    # Linha de estatísticas
    story.append(ColorBar(CONTENT_WIDTH, 1, BORDER_COLOR))
    story.append(Spacer(1, 4*mm))
    stats_text = (
        f'<b>{total_categories}</b> categorias  ●  '
        f'<b>{total_indicators}</b> indicadores'
    )
    story.append(Paragraph(stats_text, style_cover_stats))
    story.append(Spacer(1, 4*mm))
    story.append(ColorBar(CONTENT_WIDTH, 1, BORDER_COLOR))

    # Transição para template de conteúdo
    story.append(NextPageTemplate('content'))
    story.append(PageBreak())

    # ── SUMÁRIO ─────────────────────────────────────────────────────────────
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph("Sumário", style_toc_title))
    story.append(Spacer(1, 3*mm))

    for cat, inds in categories.items():
        count = len(inds)
        toc_text = (
            f'<font color="{ACCENT.hexval()}">■</font>  '
            f'{escape_xml(cat)} '
            f'<font color="{GRAY_LIGHT.hexval()}">({count} indicador{"es" if count != 1 else ""})</font>'
        )
        story.append(Paragraph(toc_text, style_toc_item))

    story.append(PageBreak())

    # ── CATEGORIAS E INDICADORES ────────────────────────────────────────────
    for cat_idx, (cat, inds) in enumerate(categories.items()):
        count = len(inds)

        # Banner da categoria
        story.append(build_category_banner(cat, count))
        story.append(Spacer(1, 3*mm))

        # Descrição da categoria
        desc_box = build_category_description(cat)
        if desc_box:
            story.append(desc_box)
            story.append(Spacer(1, 4*mm))

        # Indicadores
        for ind in inds:
            story.append(build_indicator_block(ind))

        # Quebra de página entre categorias (exceto a última)
        if cat_idx < len(categories) - 1:
            story.append(PageBreak())

    # ── GERAR ───────────────────────────────────────────────────────────────
    print(f"Gerando PDF: {output_path}")
    doc.build(story)
    print(f"Concluído! {output_path}")
    print(f"  → {doc.page} páginas geradas")


# ═══════════════════════════════════════════════════════════════════════════
# EXECUÇÃO
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # Permite passar caminhos como argumentos
    excel = sys.argv[1] if len(sys.argv) > 1 else EXCEL_INPUT
    output = sys.argv[2] if len(sys.argv) > 2 else PDF_OUTPUT

    if not os.path.exists(excel):
        sys.exit(f"Erro: arquivo não encontrado → {excel}")

    generate_pdf(excel, output)